import os
from openpyxl import load_workbook
from google.oauth2 import service_account
from googleapiclient.discovery import build
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

# Path to your service account credentials JSON file
credentials_file = 'C:/Users/lenovo/Desktop/docsProject/ops-do-ab53a8422146.json'

# Create credentials for Google Drive API
credentials = service_account.Credentials.from_service_account_file(
    credentials_file, scopes=['https://www.googleapis.com/auth/drive'])

# Build the Drive API service
drive_service = build('drive', 'v3', credentials=credentials)

# Define the path to your desktop folder
desktop_folder = os.path.expanduser("C:/Users/lenovo/Desktop/docsProject/ops_docs")
excel_file = 'C:/Users/lenovo/Desktop/docsProject/docs.xlsx'

# Load the Excel workbook
workbook = load_workbook(excel_file)
sheet = workbook.active

# Iterate over rows and columns to find photo URLs
for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column), start=2):
    driver_id = row[0].value  # driver ID is in the first column
    

    for column_index, cell in enumerate(row[1:], start=2):  # Start from the second column
        photo_file_link = cell.value
        # Find the position of '/d/' and '/view' in the link
        start_index = photo_file_link.find('/d/') + 3
        end_index = photo_file_link.find('/view')

        # Extract the file ID using slicing
        photo_file_id = photo_file_link[start_index:end_index]

        if photo_file_id:
            photo_column_header = sheet.cell(row=1, column=column_index).value  # Get the column header
            filename = f"{driver_id} - {photo_column_header}.jpg"

            # Use the Drive API service to download the file
            request = drive_service.files().get_media(fileId=photo_file_id)
            file_content = request.execute()

            # Save the file content to the desktop folder
            with open(os.path.join(desktop_folder, filename), 'wb') as f:
                f.write(file_content)

            print(f"Downloaded {filename} to desktop folder.")
    print('Files downloaded')

    # Create a dictionary to hold the file paths
    files = {
        'nationalCard': open(f'{desktop_folder}/{driver_id} - nid_front.jpg', 'rb'),
        'nationalCardBack': open(f'{desktop_folder}/{driver_id} - nid_back.jpg', 'rb'),
        'drivingCertificate': open(f'{desktop_folder}/{driver_id} - licence_front.jpg', 'rb'),
        'drivingCertificateBack': open(f'{desktop_folder}/{driver_id} - licence_back.jpg', 'rb'),
        'registerForm': open(f'{desktop_folder}/{driver_id} - ijaza_front.jpg', 'rb'),
        'registerFormBack': open(f'{desktop_folder}/{driver_id} - ijaza_back.jpg', 'rb'),
        'background': open(f'{desktop_folder}/{driver_id} - nid_front.jpg', 'rb'),
        'termsAndConditions': open(f'{desktop_folder}/{driver_id} - nid_front.jpg', 'rb'),
        'other': open(f'{desktop_folder}/{driver_id} - others.jpg', 'rb')
    }

    # Initialize Chrome webdriver
    driver = webdriver.Chrome()

    # Define the URL of the login page and the form submission endpoint
    login_url = 'https://backoffice.baly.app/auth/sign-in'
    form_submission_url = f'https://backoffice.baly.app/admin/drivers/{driver_id}/docs'

    # Navigate to the login page
    driver.get(login_url)

    # Enter your credentials and submit the login form
    username_input = driver.find_element(By.ID, 'email')
    password_input = driver.find_element(By.ID, 'password')

    # Replace 'your_username' and 'your_password' with your actual credentials
    username_input.send_keys('zahraa.jabar@baly.iq')
    password_input.send_keys('#Uu.t4j*E3Mgkz)')

    login_button = driver.find_element(By.XPATH, '//input[@type="submit"]')
    login_button.click()
    time.sleep(2)

    # Now that you are logged in, proceed to the form submission page
    driver.get(form_submission_url)
    time.sleep(1)

    # Find and upload files to the form
    for field_name, file_obj in files.items():
        file_input = driver.find_element(By.NAME, field_name)
        file_input.send_keys(file_obj.name)

        try:
            # Find and click the upload button
            upload_button = driver.find_element(By.XPATH, "//button[contains(text(), 'Upload')]")
            upload_button.click()
            time.sleep(2)
            # Print the uploading status
            print(f"{driver_id} - {field_name} Uploaded")
        except:
            # Print the uploading status
            print(f"{driver_id} - {field_name} Faild Uploading")

    # Close the webdriver
    driver.quit()

    # # List all files in the folder
    # files_in_folder = os.listdir(desktop_folder)

    # # Iterate over each file in the folder
    # for file_name in files_in_folder:
    # # Construct the full path to the file
    #     file_path = os.path.join(desktop_folder, file_name)
    
    #     # Check if the file is a regular file (not a directory)
    #     if os.path.isfile(file_path):
    #         try:
    #             # Attempt to delete the file
    #             os.remove(file_path)
    #             print(f"Deleted file: {file_path}")
    #         except Exception as e:
    #             # Handle exceptions if any
    #             print(f"Error deleting file: {file_path} - {e}")

# Close the workbook after finishing
workbook.close()
